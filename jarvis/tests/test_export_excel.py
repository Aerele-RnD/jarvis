"""Tests for export_excel's "no data" guard.

A workbook with no columns, no data rows, or only blank cells is empty in
every sense — export_excel must raise NoDataError (which the agent relays as
"No data to prepare for Excel.") rather than hand the user a blank .xlsx to
open and discover is empty.
"""
from __future__ import annotations

from unittest.mock import patch

import frappe
from frappe.tests.utils import FrappeTestCase

from jarvis.exceptions import InvalidArgumentError, NoDataError
from jarvis.tools.export_excel import export_excel


class TestExportExcelNoData(FrappeTestCase):
    def test_rejects_empty_list(self):
        with self.assertRaises(NoDataError):
            export_excel([])

    def test_rejects_list_of_empty_dicts(self):
        # No keys → no columns → a contentless workbook. Must be rejected.
        with self.assertRaises(NoDataError):
            export_excel([{}])
        with self.assertRaises(NoDataError):
            export_excel([{}, {}])

    def test_rejects_all_blank_dict_rows(self):
        # Columns exist but every value is empty → nothing to export.
        with self.assertRaises(NoDataError):
            export_excel([{"a": None, "b": None}])
        with self.assertRaises(NoDataError):
            export_excel([{"a": "", "b": "   "}])

    def test_rejects_header_only_list(self):
        with self.assertRaises(NoDataError):
            export_excel([["Name", "Qty"]])

    def test_rejects_all_blank_list_rows(self):
        with self.assertRaises(NoDataError):
            export_excel([["Name", "Qty"], ["", None]])

    def test_generates_when_there_is_real_data(self):
        # Regression guard: a row with at least one non-empty cell exports.
        with patch("frappe.utils.file_manager.save_file") as m:
            m.return_value = frappe._dict(
                file_url="/private/files/ok.xlsx", file_name="ok.xlsx", file_size=42, name="F-OK"
            )
            out = export_excel([{"a": 1, "b": None}], title="ok")
        self.assertEqual(out["file_url"], "/private/files/ok.xlsx")

    def test_generates_list_of_lists_with_data(self):
        with patch("frappe.utils.file_manager.save_file") as m:
            m.return_value = frappe._dict(
                file_url="/private/files/ll.xlsx", file_name="ll.xlsx", file_size=42, name="F-LL"
            )
            out = export_excel([["Name", "Qty"], ["Widget", 5]])
        self.assertEqual(out["file_url"], "/private/files/ll.xlsx")


def _saved_bytes(mock):
    """The content bytes export_excel handed save_file (2nd positional arg)."""
    return mock.call_args.args[1]


def _load(content):
    from io import BytesIO

    from openpyxl import load_workbook

    return load_workbook(BytesIO(content))


class TestExportExcelMultiSheet(FrappeTestCase):
    def _mock_save(self):
        m = patch("frappe.utils.file_manager.save_file").start()
        self.addCleanup(patch.stopall)
        m.return_value = frappe._dict(
            file_url="/private/files/wb.xlsx", file_name="wb.xlsx", file_size=99, name="F-WB"
        )
        return m

    def test_builds_named_tabs_in_order(self):
        m = self._mock_save()
        out = export_excel(
            title="MIS",
            sheets=[
                {"title": "Financials", "rows": [{"metric": "Revenue", "value": 100}]},
                {"title": "Customers", "rows": [["Name", "Due"], ["Acme", 50]]},
            ],
        )
        self.assertEqual(out["file_url"], "/private/files/wb.xlsx")
        wb = _load(_saved_bytes(m))
        self.assertEqual(wb.sheetnames, ["Financials", "Customers"])
        self.assertEqual(wb["Financials"]["A1"].value, "metric")   # header from dict keys
        self.assertEqual(wb["Financials"]["A2"].value, "Revenue")
        self.assertEqual(wb["Customers"]["B2"].value, 50)           # list-of-lists body

    def test_empty_tabs_are_skipped_but_the_rest_ship(self):
        m = self._mock_save()
        export_excel(
            title="w",
            sheets=[
                {"title": "HR", "rows": []},                 # empty → skipped
                {"title": "Ops", "rows": [{"a": 1}]},        # kept
                {"title": "Blank", "rows": [{"a": None}]},   # all-blank → skipped
            ],
        )
        self.assertEqual(_load(_saved_bytes(m)).sheetnames, ["Ops"])

    def test_all_empty_raises_no_data(self):
        with self.assertRaises(NoDataError):
            export_excel(sheets=[{"title": "A", "rows": []}, {"title": "B", "rows": [{}]}])

    def test_empty_sheets_list_is_invalid(self):
        with self.assertRaises(InvalidArgumentError):
            export_excel(sheets=[])

    def test_duplicate_titles_get_unique_sheet_names(self):
        m = self._mock_save()
        export_excel(
            sheets=[
                {"title": "Summary", "rows": [{"a": 1}]},
                {"title": "Summary", "rows": [{"a": 2}]},
            ],
        )
        self.assertEqual(_load(_saved_bytes(m)).sheetnames, ["Summary", "Summary-2"])
