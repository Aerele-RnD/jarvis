"""Read an attached file's contents so the agent can act on what the customer
sent — the read side of the artifact story (``export_excel`` / ``download_pdf``
are the write side).

The chat uploads attachments to Frappe's File doctype (private) and shows the
agent a ``📎 <name>`` marker plus the ``file_url``. The agent has no way to open
that binary on its own, so it would otherwise flail (try the browser tool on a
private URL, or hunt for a matching DocType). This tool fetches the File —
gated by the calling user's File read permission — and returns its contents in
a model-readable shape:

  - spreadsheets (xlsx / csv)  → rows (per sheet)
  - pdf                        → extracted text (per page)
  - text-like (txt/md/json/html/log/yaml/xml/tsv) → decoded text
  - images (png/jpg/…)         → metadata (dimensions/format); pixel content
                                 needs a vision input, not a text tool, so we
                                 say so instead of dumping useless base64.

Permission contract: the calling user must have **read** permission on the
File (Frappe's File ``has_permission`` walks the attached-doc perms for private
files). A user who can't read the file gets a clean PermissionDeniedError.
"""
import csv
import io

import frappe

from jarvis.exceptions import InvalidArgumentError, PermissionDeniedError

_TABLE_EXT = {"xlsx", "csv", "tsv"}
_PDF_EXT = {"pdf"}
_IMAGE_EXT = {"png", "jpg", "jpeg", "gif", "webp", "bmp"}
_TEXT_EXT = {"txt", "md", "markdown", "json", "html", "htm", "log", "yaml", "yml", "xml", "csv", "tsv", "svg"}

_MAX_ROWS = 1000
_MAX_CHARS = 40000

# OCR is expensive and a crafted PDF/image can DoS a worker - cap the work.
_OCR_MAX_PAGES = 30
_OCR_MAX_PIXELS = 25_000_000
_OCR_MAX_BYTES = 25 * 1024 * 1024
_OCR_TIMEOUT = 25


def read_file(
    file_url: str | None = None,
    filename: str | None = None,
    sheet: str | None = None,
    max_rows: int = 500,
    max_chars: int = 20000,
    ocr: bool | None = None,
) -> dict:
    """Read an attached file and return its contents.

    Identify the file by ``file_url`` (preferred — exact) or ``filename`` (the
    name shown in the ``📎`` marker; resolves to the most recent matching File
    the user can read). ``sheet`` selects one worksheet by name for xlsx;
    omitted returns the first sheet. ``max_rows`` / ``max_chars`` bound the
    payload so a huge file doesn't blow the turn's context. ``ocr`` controls
    optical character recognition: None (auto - OCR a scanned PDF only when text
    extraction comes back empty), True (also OCR image files), False (never OCR).
    OCR needs the tesseract engine installed; when absent it degrades to a note.
    """
    fdoc = _resolve_file(file_url, filename)
    if not frappe.has_permission("File", "read", doc=fdoc.name):
        raise PermissionDeniedError(f"no read permission on file {fdoc.file_name!r}")

    ext = (fdoc.file_name or "").rsplit(".", 1)[-1].lower() if "." in (fdoc.file_name or "") else ""
    content = fdoc.get_content()
    if isinstance(content, str):
        content = content.encode("utf-8", "replace")

    max_rows = min(int(max_rows or 500), _MAX_ROWS)
    max_chars = min(int(max_chars or 20000), _MAX_CHARS)

    base = {"filename": fdoc.file_name, "file_url": fdoc.file_url, "size_bytes": len(content)}

    if ext in _PDF_EXT:
        return {**base, **_read_pdf(content, max_chars, ocr)}
    if ext == "xlsx":
        return {**base, **_read_xlsx(content, sheet, max_rows)}
    if ext in ("csv", "tsv"):
        return {**base, **_read_csv(content, ext, max_rows)}
    if ext in _IMAGE_EXT:
        return {**base, **_read_image(content, ext, ocr)}
    if ext in _TEXT_EXT or _looks_text(content):
        return {**base, **_read_text(content, max_chars)}
    if ext == "xls":
        raise InvalidArgumentError("legacy .xls isn't supported; re-save as .xlsx and retry")
    return {**base, "kind": "binary", "note": f"unsupported file type {ext!r}; cannot extract text"}


def _resolve_file(file_url: str | None, filename: str | None):
    if file_url:
        name = frappe.db.get_value("File", {"file_url": file_url}, "name")
        if not name:
            raise InvalidArgumentError(f"no file found for url {file_url!r}")
        return frappe.get_doc("File", name)
    if filename:
        rows = frappe.get_all(
            "File", filters={"file_name": filename}, fields=["name"],
            order_by="creation desc", limit=1,
        ) or frappe.get_all(
            "File", filters={"file_name": ["like", f"%{filename}%"]}, fields=["name"],
            order_by="creation desc", limit=1,
        )
        if not rows:
            raise InvalidArgumentError(f"no file found named {filename!r}")
        return frappe.get_doc("File", rows[0].name)
    raise InvalidArgumentError("pass file_url or filename to identify the attachment")


def _read_pdf(content: bytes, max_chars: int, ocr: bool | None = None) -> dict:
    from pypdf import PdfReader

    reader = PdfReader(io.BytesIO(content))
    pages = []
    total = 0
    for page in reader.pages:
        try:
            txt = page.extract_text() or ""
        except Exception:
            txt = ""
        pages.append(txt)
        total += len(txt)
        if total >= max_chars:
            break
    text = "\n\n".join(pages)
    page_count = len(reader.pages)
    used_ocr = False

    # Scanned PDFs carry no embedded text, so pypdf returns ~nothing. Fall back
    # to OCR when the extraction is essentially empty, the caller didn't disable
    # it (ocr is not False), and a tesseract engine is actually available.
    if (ocr is not False and len(content) <= _OCR_MAX_BYTES
            and len(text.strip()) < 10 * max(1, page_count) and _ocr_available()):
        try:
            ocr_text = _ocr_pdf(content, max_chars)
        except Exception:
            ocr_text = ""
        if len(ocr_text.strip()) > len(text.strip()):
            text, used_ocr = ocr_text, True

    result = {
        "kind": "pdf",
        "page_count": page_count,
        "text": text[:max_chars],
        "truncated": len(text) > max_chars,
        "ocr": used_ocr,
    }
    if not text.strip() and not used_ocr:
        why = "disabled for this call" if ocr is False else "not available on this server"
        result["note"] = f"No extractable text; looks like a scanned PDF and OCR is {why}."
    return result


def _read_xlsx(content: bytes, sheet: str | None, max_rows: int) -> dict:
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    names = wb.sheetnames
    targets = [sheet] if sheet and sheet in names else ([names[0]] if names else [])
    out_sheets = []
    for sname in targets:
        ws = wb[sname]
        rows = []
        for r in ws.iter_rows(values_only=True):
            rows.append([_cell(c) for c in r])
            if len(rows) >= max_rows:
                break
        out_sheets.append({"name": sname, "rows": rows, "row_count": len(rows),
                           "truncated": ws.max_row is not None and ws.max_row > len(rows)})
    wb.close()
    return {"kind": "table", "format": "xlsx", "sheet_names": names, "sheets": out_sheets}


def _read_csv(content: bytes, ext: str, max_rows: int) -> dict:
    text = content.decode("utf-8-sig", "replace")
    delim = "\t" if ext == "tsv" else ","
    rows = []
    for row in csv.reader(io.StringIO(text), delimiter=delim):
        rows.append(row)
        if len(rows) >= max_rows:
            break
    return {"kind": "table", "format": ext,
            "sheets": [{"name": "Sheet1", "rows": rows, "row_count": len(rows)}]}


def _read_text(content: bytes, max_chars: int) -> dict:
    text = content.decode("utf-8", "replace")
    return {"kind": "text", "text": text[:max_chars], "truncated": len(text) > max_chars}


def _read_image(content: bytes, ext: str, ocr: bool | None = None) -> dict:
    out = {"kind": "image", "format": ext}
    im = None
    try:
        from PIL import Image

        Image.MAX_IMAGE_PIXELS = _OCR_MAX_PIXELS  # PIL raises past this (decompression bomb)
        im = Image.open(io.BytesIO(content))
        out["width"], out["height"] = im.size
        out["mode"] = im.mode
    except Exception:
        im = None

    # Image OCR is opt-in (ocr=True): it's costly and most images aren't text.
    if ocr is True and im is not None and len(content) <= _OCR_MAX_BYTES and _ocr_available():
        try:
            out["text"] = _ocr_pil(im)
            out["ocr"] = True
            out["note"] = (
                "Text extracted via OCR. Visual/diagram content (not text) still "
                "needs a vision model, not this tool."
            )
            return out
        except Exception:
            pass
    out["note"] = (
        "Image metadata only. Pixel/visual content can't be extracted as text by a "
        "tool - to analyse what the image shows it needs a vision model as an image "
        "input, not read here. Pass ocr=true to OCR any text in the image."
    )
    return out


_OCR_OK = None


def _ocr_available() -> bool:
    """True iff pytesseract + a working tesseract binary are present (memoized)."""
    global _OCR_OK
    if _OCR_OK is None:
        try:
            import pytesseract

            pytesseract.get_tesseract_version()
            _OCR_OK = True
        except Exception:
            _OCR_OK = False
    return _OCR_OK


def _ocr_pil(im) -> str:
    import pytesseract

    return pytesseract.image_to_string(im, timeout=_OCR_TIMEOUT) or ""


def _ocr_pdf(content: bytes, max_chars: int) -> str:
    """Rasterize each page and OCR it with tesseract.

    Uses pypdfium2 (Apache/BSD, permissive) for rendering - NOT PyMuPDF, which
    is AGPL and would impose copyleft on this proprietary app.
    """
    import pypdfium2 as pdfium
    import pytesseract

    out = []
    total = 0
    pdf = pdfium.PdfDocument(content)
    try:
        n = min(len(pdf), _OCR_MAX_PAGES)  # cap page count
        for i in range(n):
            page = pdf[i]
            scale = 200 / 72  # ~200 DPI
            try:
                w, h = page.get_size()  # points (1/72 inch)
                if w and h and (w * scale) * (h * scale) > _OCR_MAX_PIXELS:
                    scale = (_OCR_MAX_PIXELS / (w * h)) ** 0.5  # clamp to pixel budget
            except Exception:
                pass
            txt = pytesseract.image_to_string(
                page.render(scale=scale).to_pil(), timeout=_OCR_TIMEOUT
            ) or ""
            out.append(txt)
            total += len(txt)
            if total >= max_chars:
                break
    finally:
        pdf.close()
    return "\n\n".join(out)


def _looks_text(content: bytes) -> bool:
    sample = content[:2048]
    if not sample:
        return True
    if b"\x00" in sample:
        return False
    try:
        sample.decode("utf-8")
        return True
    except UnicodeDecodeError:
        return False


def _cell(v):
    import datetime
    import decimal

    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date, datetime.time)):
        return v.isoformat()
    if isinstance(v, decimal.Decimal):
        return float(v)
    return v
